import pandas as pd
import csv
import openpyxl
import requests
import io
from skyfield.api import N, S, E, W, wgs84
from skyfield.api import load
import os
import time
import msvcrt as m

exit = True
try:
    while exit == True:
        cwd = os.getcwd()
        pd.set_option('display.max_columns', None)  # or 1000
        pd.set_option('display.max_rows', None)  # or 1000
        pd.set_option('display.max_colwidth', None)  # or 199


        df2 = pd.read_excel('Gyro_Sunshot.xlsx', names=['N.B', 'Date', 'Day', 'Month', 'Year', 'Unnamed: 5', 'Unnamed: 6', 'Unnamed: 7', 'Unnamed: 8', 'Unnamed: 9', 'Unnamed: 10'], header=0)

        df1 = df2.drop(['N.B', 'Date', 'Day', 'Month', 'Year'], axis=1)
        df = df1.drop([0, 1, 2, 3, 4], axis=0)
        df.head()
        df = df.reset_index(drop=True)

        df = df.rename(columns={'Unnamed: 5': "S/N",
                                'Unnamed: 6': "Time",
                                'Unnamed: 7': "Obs_Sun_D",
                                'Unnamed: 8': "Obs_Sun_M",
                                'Unnamed: 9': "Obs_Sun_S",
                                'Unnamed: 10': "Obs_Gyro"})

        columns = list(df)
        #for i in range(len(columns)):
        for i in range(len(df[columns[2]])):
            if (int(df[columns[2]][i])) >= 360:
                print(f'Check Degree value {i + 1}\nShould not be greater than 360')
                exit = False
                break
            elif (int(df[columns[3]][i])) >= 60:
                print(f'Check Minute value {i + 1}\nShould not be greater than 60')
                exit = False
                break
            elif (int(df[columns[4]][i])) >= 60:
                print(f'Check Seconds value {i + 1}\nShould not be greater than 60')
                exit = False
                break
        if exit == False:
            break
        d1 = {'S/N':  df['S/N'],
              'Obs_Sun_D': df['Obs_Sun_D'],
              'Obs_Sun_M': df['Obs_Sun_M'],
              'Obs_Sun_S': df['Obs_Sun_S']
            }

        df1 = pd.DataFrame(data=d1)

        df1["Obs_Sun_decimal"] = df1[["Obs_Sun_D", "Obs_Sun_M", "Obs_Sun_S"]].apply(lambda row: float(row.values[0]) + float(row.values[1])/60 + float(row.values[2])/3600, axis=1)
        df1.head()

        df['Obs_Sun_decimal'] = df1['Obs_Sun_decimal']
        df

        wb = openpyxl.load_workbook('Gyro_Sunshot.xlsx')
        worksheet = wb.active
        date = f'{int(worksheet.cell(row = 2, column = 5).value)}-{int(worksheet.cell(row = 2, column = 4).value)}-{int(worksheet.cell(row = 2, column = 3).value)}'
        date_list = []
        d2 = {'Date': [date] * (df.shape[0])}

        df2 = pd.DataFrame(data=d2)
        df2['Time'] = df.Time

        df2['Datetime_UTC'] = pd.to_datetime(df2['Date'].astype(str) + ' ' + df2['Time'].astype(str))
        df2

        df['Obs_Sun_decimal'] = df1['Obs_Sun_decimal']
        df['Datetime_UTC'] = df2['Datetime_UTC']
        col_order = ['S/N', 'Datetime_UTC', 'Obs_Sun_D', 'Obs_Sun_M', 'Obs_Sun_S', 'Obs_Sun_decimal', 'Obs_Gyro']
        df.drop(['Time'], axis=1)
        df = df.reindex(columns=col_order)
        df

        df2['Converted_Datetime'] = df['Datetime_UTC'].dt.strftime('%Y,%m,%d,%H,%M,%S')
        cdt = df2.Converted_Datetime

        lat = float(worksheet.cell(row = 4, column = 2).value) + (float(worksheet.cell(row = 4, column = 3).value) / 60) + (float(worksheet.cell(4, 4).value) / 3600)
        lon = float(worksheet.cell(5, 2).value) + (float(worksheet.cell(row = 5, column = 3).value) / 60) + (float(worksheet.cell(row = 5, column = 4).value) / 3600)


        # def get_elevation(lat, lon):
            # query = ('https://api.open-elevation.com/api/v1/lookup'
                     # f'?locations={lat},{lon}')
            # r = requests.get(query).json()
            # elevation = pd.json_normalize(r, 'results')['elevation'].values[0]
            # return elevation


        # height_m = get_elevation(lat, lon)
        height_m = 0

        data = df2.to_string
        df2 = pd.read_csv(io.StringIO(str(data)), sep=' \s+', engine='python')
        df2['Datetime_UTC'] = pd.to_datetime(df2['Datetime_UTC'].replace('>', '', regex=True), format='%Y,%m,%d,%H,%M,%S')

        df2['Converted_Datetime'] = df2['Datetime_UTC'].dt.strftime('%Y,%m,%d,%H,%M,%S')


        def calc_az(tutc):
            yr = int(tutc.split(',')[0])
            mo = int(tutc.split(',')[1])
            da = int(tutc.split(',')[2])
            hr = int(tutc.split(',')[3])
            mi = int(tutc.split(',')[4])
            se = int(tutc.split(',')[5])

            ts = load.timescale()
            # t = ts.utc(2019, 12, 31)
            planets = load('de421.bsp')

            earth, sun = planets['earth'], planets['sun']

            earth = planets['earth']
            position = earth + wgs84.latlon(lat * N, lon * E, elevation_m=height_m)
            astro = position.at(ts.utc(yr, mo, da, hr, mi, se)).observe(sun)

            app = astro.apparent()
            alt, az, distance = app.altaz()
            dt_utc = df2['Datetime_UTC']
            return '{:.3f}'.format(az.degrees)


        az_df = df2['Converted_Datetime'].apply(calc_az)
        Azimuth = az_df.values.tolist()
        Azimuth = [float(x) for x in Azimuth]
        df2['Azimuth'] = Azimuth

        Output = df
        Output['Azimuth'] = df2.Azimuth
        Output = Output.drop(['Obs_Sun_D', 'Obs_Sun_M', 'Obs_Sun_S'], axis=1)
        Output['Final_Azimuth'] = (360 - Output.Obs_Sun_decimal) + Output.Azimuth
        Output['C - O'] = Output.Final_Azimuth - Output.Obs_Gyro.astype(float)
        Output['Mean_C-O'] = Output['C - O'].mean()


        column_reorder = ["S/N", "Datetime_UTC", "Obs_Sun_decimal", "Obs_Gyro", "Azimuth", "Final_Azimuth", "C - O", 'Mean_C-O']
        Output = Output.reindex(columns=column_reorder)
        Output.to_excel('Sunshot_Output.xlsx', index=False)
        Output
        print(Output)
        print("\nCheck C - O value in Sunshot_Output file\n \nPress any key to exit")
        break
except ValueError:
    print("Check that Date, Time, Degrees, Minutes and Seconds are numeric values only")
time.sleep(2)
os.startfile('Sunshot_Output.xlsx')
print("Please wait while the output file opens up")
def wait():
    m.getch()


wait()

